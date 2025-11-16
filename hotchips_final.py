import os
import re
import glob
import sys
import time
from datetime import datetime
from urllib.parse import quote_plus
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from playwright.sync_api import sync_playwright

# ----------------------------
# Directories
# ----------------------------
OUTPUT_DIR = "extracted_data"
DUP_DIR = "duplicate_data"
DEBUG_DIR = "debug_failures"
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(DUP_DIR, exist_ok=True)
os.makedirs(DEBUG_DIR, exist_ok=True)

# standard link font for excel
LINK_FONT = Font(color="0000FF", underline="single")

# ----------------------------
# Helpers
# ----------------------------
def now_ts():
    return datetime.now().strftime("%Y_%m_%d_%H_%M_%S")

def validate_phone(raw_text):
    if not raw_text:
        return "NA"
    s = str(raw_text)
    digits = "".join(re.findall(r'\d+', s))
    if 10 <= len(digits) <= 13:
        return digits
    return "NA"

def extract_coords_from_url(url):
    try:
        m = re.search(r'@(-?\d+\.\d+),(-?\d+\.\d+)', url)
        if m:
            return f"{m.group(1)},{m.group(2)}"
        m2 = re.search(r'!3d(-?\d+\.\d+)!4d(-?\d+\.\d+)', url)
        if m2:
            return f"{m2.group(1)},{m2.group(2)}"
    except:
        pass
    return None

# ----------------------------
# Excel / duplicate helpers (unchanged behaviour except hyperlinking)
# ----------------------------
def base_dup_path():
    return os.path.join(DUP_DIR, "duplicated.xlsx")

def find_latest_timestamped_dup():
    pattern = os.path.join(DUP_DIR, "duplicated_*.xlsx")
    files = glob.glob(pattern)
    if not files:
        return None
    files.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return files[0]

def is_workbook_empty(path):
    try:
        wb = load_workbook(path)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        if not rows:
            return True
        for r in rows:
            if any(cell is not None and str(cell).strip() != "" for cell in r):
                return False
        return True
    except Exception:
        return False

def read_entries_from_dup(path):
    entries = set()
    try:
        wb = load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            shop = (row[1] or "").strip().lower()
            location = (row[3] or "").strip()
            entries.add((shop, location))
        wb.close()
    except Exception as e:
        print(f"Warning: couldn't read duplicates file {path}: {e}")
    return entries

def save_timestamped_dup(rows):
    """
    rows: iterable of rows [[date, shop_name, phone, area_location, google_maps_of_the_area], ...]
    this function now makes the last column a clickable hyperlink when possible
    """
    timestamp = now_ts()
    path = os.path.join(DUP_DIR, f"duplicated_{timestamp}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(["date", "shop_name", "phone_number", "area_location", "google_maps_of_the_area"])
    for r in rows:
        ws.append(r)
        # turn last appended cell into hyperlink if it looks like a URL
        try:
            row_idx = ws.max_row
            url_val = (r[4] or "").strip()
            if url_val and url_val.upper() != "NA" and url_val.lower().startswith("http"):
                cell = ws.cell(row=row_idx, column=5)
                cell.value = url_val
                cell.hyperlink = url_val
                cell.font = LINK_FONT
        except Exception:
            pass
    wb.save(path)
    wb.close()
    return path

def append_to_and_update_timestamp(existing_path, new_rows):
    existing_set = read_entries_from_dup(existing_path)
    filtered_rows = []
    for r in new_rows:
        shop_key = (r[1] or "").strip().lower()
        loc_key = (r[3] or "").strip()
        if (shop_key, loc_key) not in existing_set:
            filtered_rows.append(r)
    try:
        wb_existing = load_workbook(existing_path)
        ws_existing = wb_existing.active
    except Exception as e:
        print(f"Warning: couldn't open existing duplicated file: {e}. Creating fresh file.")
        return save_timestamped_dup(new_rows), len(new_rows)
    appended_count = 0
    for fr in filtered_rows:
        ws_existing.append(fr)
        # hyperlink last column if URL
        try:
            row_idx = ws_existing.max_row
            url_val = (fr[4] or "").strip()
            if url_val and url_val.upper() != "NA" and url_val.lower().startswith("http"):
                cell = ws_existing.cell(row=row_idx, column=5)
                cell.value = url_val
                cell.hyperlink = url_val
                cell.font = LINK_FONT
        except Exception:
            pass
        appended_count += 1
    new_path = os.path.join(DUP_DIR, f"duplicated_{now_ts()}.xlsx")
    wb_existing.save(new_path)
    wb_existing.close()
    try:
        os.remove(existing_path)
    except Exception:
        pass
    return new_path, appended_count

def load_all_previous_entries():
    entries = set()
    for file in os.listdir(OUTPUT_DIR):
        if file.startswith("main_") and file.endswith(".xlsx"):
            path = os.path.join(OUTPUT_DIR, file)
            try:
                wb = load_workbook(path)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    shop = (row[1] or "").strip().lower()
                    if shop:
                        entries.add(shop)
                wb.close()
            except Exception as e:
                print(f"Warning: couldn't read {path}: {e}")
    return entries

# ----------------------------
# Validation helpers (unchanged)
# ----------------------------
def is_obviously_invalid_area(area_input):
    if not area_input or not area_input.strip():
        return True
    s = area_input.strip()
    if s.isdigit():
        return True
    letters = re.findall(r'[A-Za-z]', s)
    if not letters:
        return True
    if len("".join(letters)) < 2:
        return True
    return False

def check_area_on_maps(area_input, headless=True, timeout_ms=8000):
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=headless)
            page = browser.new_page()
            q = quote_plus(area_input)
            search_url = f"https://www.google.com/maps/search/{q}"
            try:
                page.goto(search_url, timeout=timeout_ms)
            except Exception:
                page.goto("https://www.google.com/maps")
                try:
                    page.fill("input#searchboxinput", area_input)
                except Exception:
                    page.evaluate("document.querySelector('input#searchboxinput').value = arguments[0];", area_input)
                page.keyboard.press("Enter")
            page.wait_for_timeout(1200)
            try:
                cards = page.locator("div.Nv2PK")
                if cards.count() and cards.count() > 0:
                    browser.close()
                    return True
            except Exception:
                pass
            try:
                content = page.content().lower()
                if "no results found" in content or "did not match" in content:
                    browser.close()
                    return False
            except Exception:
                pass
            url = page.url.lower()
            if "/search/" in url or "/place/" in url:
                browser.close()
                return True
            browser.close()
    except Exception as e:
        print(f"Warning: area validation step error: {e}")
        return False
    return False

def get_valid_area_from_user(max_attempts=2):
    attempt = 0
    while attempt < max_attempts:
        area_input = input("Enter area (spelling can be wrong, script will correct): ").strip()
        attempt += 1
        if is_obviously_invalid_area(area_input):
            print("Please enter a valid place name. Try again.")
            if attempt >= max_attempts:
                print("Unable to locate the area on Google Maps. Exiting.")
                sys.exit(1)
            continue
        if check_area_on_maps(area_input, headless=False):
            return area_input
        else:
            if attempt < max_attempts:
                print(f"The place '{area_input}' does not appear to exist in Bangalore.")
                print("Enter the correct area name within Bangalore:")
                continue
            print("Unable to locate the area. Exiting.")
            sys.exit(1)
    print("Area validation failed. Exiting.")
    sys.exit(1)

# ----------------------------
# Fallback address helper (keeps your previous selectors)
# ----------------------------
def extract_shop_address(page):
    candidates = [
        'button[data-item-id^="address:"]',
        'button[data-item-id="address"]',
        'button[aria-label^="Address"]',
        'div.Yr7JMd-pane-hSRGPd',
        'div.IiD88e',
        'div[data-section-id="ad"]',
    ]
    for sel in candidates:
        try:
            locator = page.locator(sel).first
            if locator:
                txt = locator.text_content()
                if txt and txt.strip():
                    return txt.strip()
        except Exception:
            pass
    try:
        content_list = page.locator('div.section-hero-header-title').all_text_contents()
        if content_list:
            return " ".join([c.strip() for c in content_list if c.strip()])
    except Exception:
        pass
    try:
        return page.url.strip()
    except:
        return "NA"

# ----------------------------
# Share link helper (unchanged)
# ----------------------------
def extract_share_link_from_dialog(page):
    try:
        share_selectors = [
            'button[aria-label^="Share"]',
            'button[aria-label="Share"]',
            'button[jsaction^="pane.share"]',
            'button[aria-label*="share"]',
            'div[aria-label="Share"] button',
        ]
        clicked = False
        for sel in share_selectors:
            try:
                btn = page.locator(sel).first
                if btn and btn.is_visible():
                    btn.click()
                    clicked = True
                    break
            except Exception:
                continue
        if not clicked:
            return "NA"
        page.wait_for_timeout(900)
        try:
            dialog = page.locator('div[role="dialog"]').first
            dlg_text = dialog.inner_text()
        except Exception:
            dlg_text = page.content()
        m = re.search(r'(https?://\S+)', dlg_text)
        if m:
            link = m.group(1).strip().rstrip(')"\'')
            try:
                page.keyboard.press("Escape")
            except:
                pass
            return link
        try:
            inp = dialog.locator('input').first
            if inp:
                try:
                    val = inp.get_attribute('value') or inp.input_value()
                    if val and val.startswith('http'):
                        try:
                            page.keyboard.press("Escape")
                        except:
                            pass
                        return val
                except Exception:
                    pass
        except Exception:
            pass
        try:
            page.keyboard.press("Escape")
        except:
            pass
    except Exception:
        pass
    return "NA"

# ----------------------------
# MAIN: improved extraction + exact place URL feature
# ----------------------------
def scrape_hot_chips(area, count_needed):
    historical_names = load_all_previous_entries()

    timestamp = now_ts()
    MAIN_FILE = os.path.join(OUTPUT_DIR, f"main_{timestamp}.xlsx")
    BASE_DUP = base_dup_path()

    wb_main = Workbook()
    ws_main = wb_main.active
    ws_main.append(["date", "shop_name", "phone_number", "area_location", "google_maps_of_the_area"])
    wb_main.save(MAIN_FILE)
    wb_main.close()

    dup_entries = []
    dup_seen = set()
    saved = 0
    seen_this_run = set()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        query = quote_plus(f"hot chips in {area}, Bangalore")
        search_url = f"https://www.google.com/maps/search/{query}"
        try:
            page.goto(search_url)
        except Exception:
            page.goto("https://www.google.com/maps")
            try:
                page.fill("input#searchboxinput", f"hot chips in {area}, Bangalore")
            except Exception:
                page.evaluate("document.querySelector('input#searchboxinput').value = arguments[0];", f"hot chips in {area}, Bangalore")
            page.keyboard.press("Enter")

        page.wait_for_timeout(1400)

        empty_retries = 0
        empty_retries_max = 6

        while saved < count_needed:
            try:
                total_cards = page.locator("div.Nv2PK").count()
            except:
                total_cards = 0

            if total_cards == 0:
                empty_retries += 1
                if empty_retries <= empty_retries_max:
                    try:
                        page.evaluate("window.scrollBy(0, 800);")
                    except:
                        pass
                    page.wait_for_timeout(1100)
                    try:
                        total_cards = page.locator("div.Nv2PK").count()
                    except:
                        total_cards = 0
                    if total_cards == 0:
                        try:
                            page.reload()
                        except:
                            pass
                        page.wait_for_timeout(1100)
                        try:
                            total_cards = page.locator("div.Nv2PK").count()
                        except:
                            total_cards = 0
                else:
                    print("No shop cards found after retries — stopping.")
                    break
            else:
                empty_retries = 0

            for idx in range(total_cards):
                if saved >= count_needed:
                    break
                try:
                    shop_el = page.locator("div.Nv2PK").nth(idx)

                    # quick name from card (fast)
                    name = "N/A"
                    try:
                        name = shop_el.locator("div.qBF1Pd").inner_text().strip()
                    except:
                        try:
                            txt = shop_el.inner_text().strip()
                            name = txt.splitlines()[0].strip() if txt else "N/A"
                        except:
                            name = "N/A"
                    norm_name = (name or "N/A").strip().lower()
                    if norm_name in seen_this_run:
                        continue

                    # click to open pane
                    try:
                        shop_el.click()
                    except:
                        try:
                            page.evaluate("arguments[0].click();", shop_el)
                        except:
                            pass

                    # Wait for important nodes; slightly generous to avoid missing lazy nodes
                    try:
                        page.wait_for_selector(
                            'div.section-hero-header-title, a[href^="tel:"], button[data-item-id^="phone:"], button[aria-label*="Copy address"], div[role="dialog"]',
                            timeout=2000
                        )
                    except:
                        page.wait_for_timeout(900)

                    # single evaluate harvest of many candidates
                    try:
                        harvested = page.evaluate(
                            """() => {
                                const out = { address: '', pageUrl: window.location.href || '', telCandidates: [], shareText: '', paneText: '' };
                                const addrSelectors = [
                                    'button[data-item-id^=\"address:\"]',
                                    'button[data-item-id=\"address\"]',
                                    'button[aria-label^=\"Address\"]',
                                    'div.Yr7JMd-pane-hSRGPd',
                                    'div.IiD88e',
                                    'div[data-section-id=\"ad"]',
                                    'button[aria-label*=\"Copy address\"]',
                                    'div.section-hero-header-title'
                                ];
                                for (const sel of addrSelectors) {
                                    try {
                                        const el = document.querySelector(sel);
                                        if (el && el.innerText && el.innerText.trim().length>0) {
                                            out.address = el.innerText.trim();
                                            break;
                                        }
                                    } catch(e){}
                                }
                                try {
                                    const pane = document.querySelector('div.section-hero-text-content') || document.querySelector('div.section-hero-header-title');
                                    if (pane) out.paneText = (pane.innerText || '').trim();
                                } catch(e){}
                                try {
                                    const tels = Array.from(document.querySelectorAll('a[href^=\"tel:\"]'));
                                    for (const t of tels) out.telCandidates.push(t.getAttribute('href') || t.innerText || '');
                                } catch(e){}
                                try {
                                    const phoneBtns = Array.from(document.querySelectorAll('button[data-item-id^=\"phone:\"], button[aria-label*=\"Call\"], button[aria-label*=\"Phone\"]'));
                                    for (const b of phoneBtns) out.telCandidates.push(b.innerText || b.getAttribute('aria-label') || '');
                                } catch(e){}
                                try {
                                    const dlg = document.querySelector('div[role=\"dialog\"]');
                                    if (dlg) out.shareText = (dlg.innerText || '').trim();
                                } catch(e){}
                                return out;
                            }"""
                        )
                    except Exception:
                        harvested = {"address": "", "pageUrl": page.url, "telCandidates": [], "shareText": "", "paneText": ""}

                    address_text = (harvested.get("address") or "").strip()
                    current_url = (harvested.get("pageUrl") or page.url or "").strip()
                    tel_candidates = harvested.get("telCandidates") or []
                    share_dialog_text = (harvested.get("shareText") or "").strip()
                    pane_text = (harvested.get("paneText") or "").strip()

                    # Normalize phone
                    phone = "NA"
                    for cand in tel_candidates:
                        ph = validate_phone(cand)
                        if ph != "NA":
                            phone = ph
                            break
                    if phone == "NA" and share_dialog_text:
                        phone = validate_phone(share_dialog_text)
                    if phone == "NA":
                        phone = validate_phone(pane_text)

                    # --- NEW: try to get exact place URL (prefer current_url if already place/coords) ---
                    exact_place_url = ""
                    # prefer current_url if it looks like a place url or contains coords
                    if current_url and ("/place/" in current_url or "@"+"" in current_url or re.search(r'!3d-?\d', current_url)):
                        exact_place_url = current_url
                    else:
                        # try to extract coords from current_url first
                        ccoords = extract_coords_from_url(current_url)
                        if ccoords:
                            exact_place_url = current_url
                        else:
                            # fallback: try share dialog extraction (only if we don't already have a place URL)
                            try:
                                share_link = extract_share_link_from_dialog(page)
                                if share_link and share_link != "NA":
                                    exact_place_url = share_link
                                else:
                                    exact_place_url = current_url or "NA"
                            except:
                                exact_place_url = current_url or "NA"

                    # coords field (as before) - keep for backwards compat
                    coords = extract_coords_from_url(exact_place_url) or extract_coords_from_url(current_url)
                    google_maps_loc = exact_place_url if exact_place_url else (current_url or "NA")

                    # fallback for address (slower) only if empty
                    if not address_text:
                        try:
                            fallback = extract_shop_address(page)
                            if fallback and fallback.strip():
                                address_text = fallback
                        except:
                            pass

                    # If still missing phone or address — save debug screenshot
                    need_debug = (not phone or phone == "NA") or (not address_text or address_text.strip() == "")
                    debug_path = None
                    if need_debug:
                        safe_name = re.sub(r'\W+', '_', name)[:30]
                        debug_path = os.path.join(DEBUG_DIR, f"{now_ts()}_idx{idx}_{safe_name}.png")
                        try:
                            page.screenshot(path=debug_path, full_page=False)
                        except:
                            debug_path = None

                    today = datetime.now().strftime("%Y-%m-%d")

                    # record duplicate info if historically seen, but still write to main file
                    if norm_name in historical_names:
                        dup_row = (name, phone, address_text, google_maps_loc)
                        if dup_row not in dup_seen:
                            dup_entries.append([today, name, phone, address_text, google_maps_loc])
                            dup_seen.add(dup_row)
                            print(f"[DUP-RECORDED] {name}")

                    # Save to main file (append then set hyperlink on last column if URL)
                    try:
                        wb = load_workbook(MAIN_FILE)
                        ws = wb.active
                    except Exception:
                        wb = Workbook()
                        ws = wb.active
                        ws.append(["date", "shop_name", "phone_number", "area_location", "google_maps_of_the_area"])

                    ws.append([today, name, phone, address_text, google_maps_loc])
                    # set hyperlink if google_maps_loc looks like a URL
                    try:
                        r_idx = ws.max_row
                        url_val = (google_maps_loc or "").strip()
                        if url_val and url_val.upper() != "NA" and url_val.lower().startswith("http"):
                            cell = ws.cell(row=r_idx, column=5)
                            cell.value = url_val
                            cell.hyperlink = url_val
                            cell.font = LINK_FONT
                    except Exception:
                        pass

                    wb.save(MAIN_FILE)
                    wb.close()

                    saved += 1
                    seen_this_run.add(norm_name)
                    historical_names.add(norm_name)

                    print(f"Saved: {name} | phone:{phone} | addr_present:{bool(address_text)} | loc:{google_maps_loc} | saved={saved}/{count_needed}")
                    if debug_path:
                        print(f" -> saved debug screenshot: {debug_path}")

                    if saved >= count_needed:
                        break

                except Exception as e:
                    print("Error while processing shop:", e)
                    try:
                        page.keyboard.press("Escape")
                    except:
                        pass
                    continue

            if saved >= count_needed:
                break

            # try next or load more
            try:
                next_btn = page.locator("button[aria-label='Next']").first
                if next_btn and next_btn.is_visible():
                    next_btn.click()
                    page.wait_for_timeout(1000)
                    continue
                else:
                    sc_tries = 0
                    sc_max = 4
                    loaded_more = False
                    old_count = total_cards
                    while sc_tries < sc_max and saved < count_needed:
                        try:
                            page.evaluate("window.scrollBy(0, 900);")
                        except:
                            pass
                        page.wait_for_timeout(800)
                        try:
                            new_total = page.locator("div.Nv2PK").count()
                        except:
                            new_total = old_count
                        if new_total and new_total > old_count:
                            loaded_more = True
                            break
                        sc_tries += 1
                    if loaded_more:
                        continue
                    print("No more result pages/cards available.")
                    break
            except Exception:
                print("Couldn't navigate to next page or load more results.")
                break

        browser.close()

    # duplicate: append/update behavior unchanged except hyperlinking
    latest_ts_dup = find_latest_timestamped_dup()
    base_exists = os.path.exists(BASE_DUP)
    if not dup_entries:
        if not base_exists and latest_ts_dup is None:
            wb_dup = Workbook()
            ws_dup = wb_dup.active
            ws_dup.append(["date", "shop_name", "phone_number", "area_location", "google_maps_of_the_area"])
            wb_dup.save(BASE_DUP)
            wb_dup.close()
            print(f"\n🆕 Created empty base duplicate file: {BASE_DUP}")
        else:
            print("\n⭕ No duplicates found in this run. No changes to duplicate files.")
    else:
        if latest_ts_dup:
            new_path, appended = append_to_and_update_timestamp(latest_ts_dup, dup_entries)
            if base_exists and is_workbook_empty(BASE_DUP):
                try:
                    os.remove(BASE_DUP)
                except:
                    pass
            if appended > 0:
                print(f"\n⚠️ Appended {appended} new duplicate rows and updated file: {new_path}")
            else:
                print(f"\nℹ️ No new duplicate rows (already existed). Updated file: {new_path}")
        else:
            new_path = save_timestamped_dup(dup_entries)
            if base_exists and is_workbook_empty(BASE_DUP):
                try:
                    os.remove(BASE_DUP)
                except:
                    pass
            print(f"\n⚠️ Duplicate File Created: {new_path}")

    print(f"\n🎉 Main File Created: {MAIN_FILE} (rows saved this run: {saved})")

# ----------------------------
# Entrypoint
# ----------------------------
if __name__ == "__main__":
    area = get_valid_area_from_user(max_attempts=2)

    while True:
        count_input = input("How many shops to scrape?: ").strip()
        if count_input.isdigit():
            count = int(count_input)
            break
        print("❌ Please enter a valid number (no letters or special characters). Try again.")

    scrape_hot_chips(area, count)
    print("\n✅ COMPLETED SUCCESSFULLY")
